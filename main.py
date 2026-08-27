#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
worldlibrary.ai 电子书信息采集工具

流程：
1. 启动时访问 https://worldlibrary.ai/ 获取首页完整 HTML，再从首页引用的 JS bundle 中
   解析出合集中 20 个站点的站点名称（中/英文）与跳转链接（站点 key 如 main/agriculture）
2. 界面选择站点名称（可配置开始/结束年份，按年逐次请求），点击"开始执行"后：
   - 调用 https://worldlibrary.ai/wl-api/search/search（POST）翻页采集，
     每页固定最多 20 条，每次调用接口后 sleep 间隔秒；请求体带 document_type=book
     过滤（服务端只返回电子书），可选 publication_year_from/to 按出版年份过滤
   - 接口限制 from 最大 9980（单个排序方向只够到 10000 条），超量时分两级突破：
     1) 正序取前 10000 条 + 倒序取后 10000 条拼接，单个查询可覆盖到 20000 条
     2) 仍超 20000 条时，按 facet 聚合（language 等）切成互斥子分区逐个递归；
        分区维度需通过可信度检查（无漏桶、非多值字段）且比直接拼接更划算才采用
     覆盖不到的部分会在日志中如实报告数量（接口能力所限，无法做到 100%）
   - 逐本调用 https://worldlibrary.ai/wl-api/ai/book-translate 将英文书名翻译为中文
     （标题已含中文则跳过翻译，翻译失败回退原文）
   - 输出：每个站点一个"当前"Excel（output/worldlibrary_{站点key}.xlsx），
     超过 MAX_ROWS_PER_FILE 行自动归档为 worldlibrary_{key}_{序号}.xlsx 另开新文件；
     跨文件去重用纯文本 worldlibrary_{key}_ids.txt（每行一个书籍ID，非数据库），
     启动时流式读入去重（txt 丢失时会自动扫描 Excel 重建）；每累计 500 条落盘一次
3. 结束执行按钮可随时中断

说明：
- 书籍封面链接按官网 JS 中的规律由书籍ID直接拼接（官网本身也不校验是否存在）
- 搜索接口有频率限制（突发约 10 次即被拒），由全局节流器统一配速；触发后
  必须完全静默一段时间才能恢复——被拒的请求同样计入窗口，边等边试只会
  不断延长惩罚（实测每 5 秒探一次，124 秒都没恢复；完全静默 30 秒即恢复）
- 所有请求走同一个 keep-alive 会话，避免每本书都重新握手（那样既慢十倍，
  又容易被中途掐断报 SSL: UNEXPECTED_EOF_WHILE_READING）
"""

import glob
import math
import os
import re
import subprocess
import sys
import threading
import time
import tkinter as tk
from collections import deque
from datetime import datetime
from queue import Empty, Queue
from tkinter import scrolledtext, ttk
from urllib.parse import urljoin

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill

# ==================== 配置区域 ====================

HOME_URL = "https://worldlibrary.ai/"
SEARCH_API = "https://worldlibrary.ai/wl-api/search/search"
TRANSLATE_API = "https://worldlibrary.ai/wl-api/ai/book-translate"
# 站点搜索页（作为请求头 Referer），站点主页（站点跳转链接）
SEARCH_PAGE_TPL = "https://worldlibrary.ai/search/{key}?keyword=&way=search-ai"
HOME_PAGE_TPL = "https://worldlibrary.ai/home/{key}"
BOOK_URL_TPL = "https://worldlibrary.ai/book/{key}/{book_id}"

PAGE_SIZE = 20          # 接口每页固定最多返回 20 条
MAX_FROM = 9980         # 接口硬限制：from 最大 9980，from=10000 起返回空
WINDOW_LIMIT = MAX_FROM + PAGE_SIZE     # 单个排序方向最多够到 10000 条
FLIP_LIMIT = WINDOW_LIMIT * 2           # 正序+倒序两趟，最多够到 20000 条
SEAM_MARGIN = 200       # 正倒序拼接处的重叠冗余条数（同名书排序不稳定，留余量防缺口）

# 分区维度：按 facet 聚合切分。这些维度单值且分桶精确（实测 language 分桶
# 计数与过滤结果完全一致、sum_other_doc_count=0），是真正互斥的分区键。
# 注意：不要用 title:[lo TO hi] 做分片——那是 ES 分词后的 token 范围查询，
# title:[a TO a] 会命中任何含独立单词 "a" 的书名（实测 [a TO z] 命中
# 19778/19784，几乎等于全量），分片严重重叠且递归不收敛。
PARTITION_DIMS = ["language", "publisher", "subject"]
FACET_SIZE = 300        # 分桶上限，要能装下一个维度的全部取值，否则分区会有缺口

# 搜索接口节流参数（令牌桶 + 轮次硬边界，全部来自压测，别凭感觉改）：
# - 实测配额约"30 次 / 3 分钟"，突发到第 11 次即被拒
# - "9 次快发 + 静默 70 秒"压测 63 次零限流、6.7s/次；60 秒周期仍有 1~2 次限流，
#   30/45 秒在第 30 次就崩，匀速 6.5s/次也会崩——必须保留轮次间的硬空窗
# - 轮内连发时令牌桶管速度，轮间靠"每 70 秒最多 9 次"的硬边界留出空窗；
#   翻页时翻译已把请求拖慢到 ~9 秒/次，硬边界几乎不拦
REQUEST_INTERVAL = 60.0 / 9      # 基准：一个令牌约 6.7 秒（轮内连续快发时用）
SEARCH_BURST = 9                 # 每轮最多发 9 次
SEARCH_BURST_PERIOD = 70.0       # 每轮时间窗口（即"9 次 + 静默 70 秒"的周期）
SEARCH_MAX_INTERVAL = 12.0       # 触发限流后补充间隔最多拉到 12 秒
SEARCH_RECOVER_AFTER = 8         # 连续顺利多少次就把间隔收回一档
TRANSLATE_INTERVAL = 0.2  # 翻译接口每次调用后 sleep 的秒数
API_TIMEOUT = 60        # 单次请求超时秒
# 搜索接口失败会丢真实数据，退避给足：8/16/32/64/128 秒，累计约 4 分钟
# （持续压测时服务端会直接 RST 连接，短退避扛不过去）
API_RETRY = 5           # 失败/限流重试次数
API_RETRY_SLEEP = 8     # 网络异常重试前等待秒（每次重试翻倍）
# 触发限流后的静默冷却：实测完全静默 30 秒即恢复，而边等边探 124 秒都没恢复，
# 所以这段时间内绝不能再发请求（节流器的 penalize 会同时清空窗口记录）
THROTTLE_SLEEP = 35     # 触发限流后首次静默秒数（随重试翻倍）
TRANSLATE_RETRY = 2         # 翻译接口重试次数（失败可回退英文原名，不值得久等）
TRANSLATE_RETRY_SLEEP = 2   # 翻译接口重试前等待秒
FLUSH_ROWS = 500        # 每累计新增该条数，将 Excel 追加落盘一次（防意外丢失）
MAX_ROWS_PER_FILE = 200000  # 单个Excel超过该行数自动归档（另开新文件），避免文件过大加载变慢

# 代理：None 表示自动探测（环境变量 → 系统代理：Windows 注册表 / macOS scutil）；可手动指定如 "http://127.0.0.1:7897"
PROXY_OVERRIDE = None

OUTPUT_DIR = "output"

# 输出列（与《书籍信息采集模板.xlsx》一致）
EXCEL_COLUMNS = ["书籍ID", "书籍名称", "作者", "出版社", "出版时间", "ISBN",
                 "页数", "书籍封面链接", "书籍链接", "SSN号", "读秀号"]

# 搜索接口请求体字段
SOURCE_FIELDS = [
    "title", "author", "language", "publisher", "publication_year", "collection",
    "document_type", "brief_summary", "standard_summary", "subject", "keyword",
    "bisac_l1", "bisac_l2", "bisac_l3", "page", "file_size_bytes",
]
FACETS = ["author", "language", "publication_year", "publisher", "subject",
          "keyword", "document_type", "collection", "bisac", "indicator"]
FILTER_DIMS = ["author", "language", "publisher", "subject", "keyword",
               "document_type", "indicator"]

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36")

# 站点列表兜底（首页/JS 解析失败时使用）：(key, 英文名, 中文名)
FALLBACK_SITES = [
    ("main", "Main Site", "主站点"),
    ("asian", "Asian Studies", "亚洲研究"),
    ("socialscience", "Social Sciences", "社会科学"),
    ("business", "Business & Economics", "商业与经济"),
    ("agriculture", "Agriculture", "农业"),
    ("humanity", "Humanity", "人文学科"),
    ("engineering", "Engineering", "工程学"),
    ("africa", "African Studies", "非洲研究"),
    ("chinesestudies", "Chinese Studies", "中国研究"),
    ("environmental", "Environmental", "环境"),
    ("history", "History", "历史"),
    ("lifescience", "Life Science", "生命科学"),
    ("middleeast", "Middles East Studies", "中东研究"),
    ("healthscience", "Health Sciences", "健康科学"),
    ("nursing", "Nursing", "护理学"),
    ("draa", "DRAA Select", "DRAA精选集"),
    ("science", "The Sciences", "科学"),
    ("asean", "ASEAN", "东盟"),
    ("defense", "Defense", "国防"),
    ("southasia", "South Asia", "南亚"),
]

# ==================== 代理探测 ====================


def _windows_system_proxy():
    """读取 Windows 系统代理（注册表 Internet Settings，Clash/v2rayN 等客户端
    开启"系统代理"后会写入）。非 Windows 或未启用时返回 None"""
    if sys.platform != "win32":
        return None
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                            r"Software\Microsoft\Windows\CurrentVersion\Internet Settings") as k:
            enabled = winreg.QueryValueEx(k, "ProxyEnable")[0]
            server = winreg.QueryValueEx(k, "ProxyServer")[0]
    except Exception:
        return None
    if not enabled or not server:
        return None
    http = https = None
    # ProxyServer 形如 "127.0.0.1:7890"，或 "http=127.0.0.1:7890;https=...;socks=..."
    for part in str(server).split(";"):
        part = part.strip()
        if not part:
            continue
        if "=" in part:
            scheme, _, addr = part.partition("=")
            scheme = scheme.strip().lower()
        else:
            scheme, addr = "", part
        addr = addr.strip()
        if not addr or scheme == "socks":   # socks 需要额外安装 PySocks，忽略
            continue
        if "://" not in addr:
            addr = "http://" + addr
        if scheme == "http" and http is None:
            http = addr
        elif scheme == "https" and https is None:
            https = addr
        elif not scheme:
            if http is None:
                http = addr
            if https is None:
                https = addr
    if http or https:
        return {"http": http or https, "https": https or http}
    return None


def detect_proxy():
    """探测可用代理：环境变量 → 系统代理（Windows 注册表 / macOS scutil）→ 直连"""
    if PROXY_OVERRIDE:
        return {"http": PROXY_OVERRIDE, "https": PROXY_OVERRIDE}
    for name in ("https_proxy", "HTTPS_PROXY", "http_proxy", "HTTP_PROXY",
                 "all_proxy", "ALL_PROXY"):
        val = os.environ.get(name)
        if val:
            return {"http": val, "https": val}
    win = _windows_system_proxy()
    if win:
        return win
    try:
        out = subprocess.run(["scutil", "--proxy"], capture_output=True,
                             text=True, timeout=5).stdout
        enabled = re.search(r"(HTTP|HTTPS)Enable\s*:\s*1", out)
        host = re.search(r"(HTTP|HTTPS)Proxy\s*:\s*(\S+)", out)
        port = re.search(r"(HTTP|HTTPS)Port\s*:\s*(\d+)", out)
        if enabled and host and port:
            url = "http://%s:%s" % (host.group(2), port.group(2))
            return {"http": url, "https": url}
    except Exception:
        pass
    return None


PROXIES = detect_proxy()

# ==================== 全局停止控制 ====================
_stop_event = threading.Event()


class TaskStoppedException(Exception):
    pass


def check_stop():
    if _stop_event.is_set():
        raise TaskStoppedException()


def interval_sleep(seconds: float):
    """可中断的 sleep：等待期间收到停止信号则立即抛出"""
    if _stop_event.wait(timeout=seconds):
        raise TaskStoppedException()


# 翻译失败累计计数（非致命，仅用于结束时汇总）
_translate_fail = 0

# ==================== 线程安全日志队列 ====================
LOG_QUEUE = Queue(maxsize=2000)


def log_print(text):
    LOG_QUEUE.put(text)


# ==================== 搜索接口节流 ====================

class RateLimiter:
    """搜索接口的全局节流器（令牌桶）。

    实测该接口是"约 30 次 / 3 分钟"的窗口配额，而不是短突发限制。压测结论：
    - 突发约 10 次即被拒；
    - 9 次快发 + 静默 60 秒，长跑 72 次零限流；静默 30/45 秒都会在第 30 次崩；
    - 平滑配速反而更差：匀速 6.5 秒/次同样在第 30 次崩，因为窗口滚不过去。
    所以按"桶容量 9、每 60 秒回满"实现令牌桶——它天然复刻批量+静默的节奏。

    另一个好处是自适应：翻页时翻译已经把搜索拖慢到约 9 秒/次，令牌桶根本不拦；
    只有分区扫描这类连续快发才会形成"快发 9 次 → 静默等桶回满"。

    触发限流时桶清空并按倍率降速（宁可慢、不可再被罚）；长期顺利逐步收回。"""

    def __init__(self, capacity: int, refill_interval: float,
                 burst_period: float, max_refill_interval: float,
                 recover_after: int = 20):
        self.capacity = capacity                    # 桶容量（即最多连发几次）
        self.base_interval = refill_interval        # 一个令牌多久补充（基准）
        self.interval = refill_interval             # 当前补充间隔
        self.max_interval = max_refill_interval
        self.recover_after = recover_after
        self.burst_period = burst_period            # 每轮 burst_period 秒最多 capacity 次
        self.tokens = float(capacity)
        self.last = time.monotonic()
        self.burst_start = time.monotonic()         # 本轮第 1 次的时刻
        self.burst_used = 0                         # 本轮已用次数
        self.ok_streak = 0
        self.lock = threading.Lock()

    def acquire(self):
        """阻塞到可以安全发下一个请求（等待期间响应停止信号）

        规则 = 令牌桶 + 轮次硬边界：
        - 每 burst_period 秒最多 capacity 次（实测 9 次/60s 是零限流的安全节奏）
        - 轮内连发则令牌耗尽，自然排到轮次之后，形成"快发 N 次 → 静默"的空窗
        - 翻页时翻译已把请求拖慢，轮次边界几乎不拦，只在连续快发时生效"""
        while True:
            with self.lock:
                now = time.monotonic()
                if self.burst_used >= self.capacity:
                    # 本轮配额用尽：等到本轮结束（下一轮从 burst_start+period 起）
                    wait = self.burst_start + self.burst_period - now
                    if wait <= 0:
                        self.burst_start = now
                        self.burst_used = 0
                    else:
                        interval_sleep(wait)
                        continue
                self.tokens = min(self.capacity, self.tokens
                                  + (now - self.last) / self.interval)
                self.last = now
                if self.tokens >= 1.0:
                    self.tokens -= 1.0
                    self.burst_used += 1
                    return
                wait = (1.0 - self.tokens) * self.interval
            interval_sleep(wait)

    def penalize(self) -> float:
        """触发限流：清空令牌、作废本轮，并按倍率拉长补充间隔（宁可慢，不可再被罚）"""
        with self.lock:
            self.ok_streak = 0
            self.tokens = 0.0
            self.burst_used = self.capacity    # 本轮直接作废，强制静默到下一轮
            self.burst_start = time.monotonic()
            self.interval = min(self.interval * 1.5, self.max_interval)
            return self.interval

    def reward(self):
        """一次顺利请求：连续顺利够多次就把补充间隔收回一档"""
        with self.lock:
            self.ok_streak += 1
            if self.ok_streak >= self.recover_after and self.interval > self.base_interval:
                self.interval = max(self.base_interval, self.interval / 1.5)
                self.ok_streak = 0


# 翻译接口限流宽松得多（实测 160 次 × 0.2s 间隔零限流），只节流搜索接口
SEARCH_LIMITER = RateLimiter(capacity=SEARCH_BURST,
                             refill_interval=REQUEST_INTERVAL,
                             burst_period=SEARCH_BURST_PERIOD,
                             max_refill_interval=SEARCH_MAX_INTERVAL,
                             recover_after=SEARCH_RECOVER_AFTER)


# ==================== HTTP 会话（连接复用） ====================
_tls = threading.local()


def get_session() -> requests.Session:
    """线程内复用的 HTTP 会话（keep-alive + 连接池）。

    早期版本每次请求都用 requests.post 新建连接，一年两万本书就是两万次 TLS
    握手；穿过本地代理时极易被中途掐断，报
    `SSL: UNEXPECTED_EOF_WHILE_READING`。实测复用连接后单次延迟从 1.06s 降到
    0.10s，握手次数降到个位数。"""
    s = getattr(_tls, "session", None)
    if s is None:
        s = requests.Session()
        if PROXIES:
            s.proxies.update(PROXIES)
        adapter = requests.adapters.HTTPAdapter(pool_connections=4, pool_maxsize=8,
                                                max_retries=0)
        s.mount("https://", adapter)
        s.mount("http://", adapter)
        _tls.session = s
    return s


def reset_session():
    """丢弃当前会话：SSL/连接异常后池里那条连接已损坏，重建比复用更稳"""
    s = getattr(_tls, "session", None)
    if s is not None:
        try:
            s.close()
        except Exception:
            pass
    _tls.session = None


# ==================== 接口调用 ====================

def build_headers(site_key: str) -> dict:
    """请求头（Referer 使用对应站点的搜索页）"""
    return {
        'User-Agent': UA,
        'Content-Type': 'application/json',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Origin': 'https://worldlibrary.ai',
        'Referer': SEARCH_PAGE_TPL.format(key=site_key),
    }


def build_body(site_key: str, query: str = "*", filters: dict = None,
               from_: int = None, sort_order: str = "asc",
               facet: list = None, size: int = PAGE_SIZE) -> dict:
    """构造搜索请求体（与官网前端一致：过滤生效时带 _type=filter）

    排序字段用 title 而非 publication_year：按年份过滤后年份是常量，排序完全
    并列、翻页顺序不稳定；title 才能给出确定的全序，正倒序两趟才能拼得上。

    facet 默认不带：翻页只要 hits，让服务端顺带算 10 个维度的聚合会明显拖慢
    每一次请求（大年份能慢到超时）。只有 facet_buckets() 需要时才显式指定。"""
    body = {
        "search_type": "keyword",
        "query": query,
        "size": size,
        "facet_size": FACET_SIZE,
        "source_fields": list(SOURCE_FIELDS),
        "facet": list(facet or []),
        "author": [], "language": [], "publisher": [], "subject": [],
        "keyword": [], "document_type": ["book"], "indicator": [],
        "sort_field": "title",
        "sort_order": sort_order,
    }
    if site_key != "main":          # 主站点不带 collection，其余站点按站点 key 过滤合集
        body["collection"] = [site_key]
    filters = filters or {}
    has_filter = True               # 请求体始终带 document_type=["book"] 过滤
    for dim in FILTER_DIMS:
        if filters.get(dim):
            body[dim] = list(filters[dim])
    if filters.get("publication_year_from") is not None or filters.get("publication_year_to") is not None:
        body["publication_year_from"] = filters.get("publication_year_from")
        body["publication_year_to"] = filters.get("publication_year_to")
    if has_filter:
        body["_type"] = "filter"
    if from_ is not None:
        body["from"] = from_
    return body


def fetch_json(url: str, payload: dict, headers: dict, timeout: int = API_TIMEOUT,
               retries: int = API_RETRY, retry_sleep: float = API_RETRY_SLEEP,
               quiet: bool = False, limiter: "RateLimiter" = None):
    """POST JSON 接口，带节流、失败重试与限流退避。成功返回 dict，失败返回 None"""
    wait = THROTTLE_SLEEP
    for attempt in range(1, retries + 1):
        check_stop()
        if limiter is not None:
            limiter.acquire()       # 阻塞到满足最小间隔和窗口配额
        try:
            res = get_session().post(url, json=payload, headers=headers, timeout=timeout)
        except requests.RequestException as e:
            # SSL/连接类异常：连接池里这条连接已废，重建会话再试
            if isinstance(e, (requests.exceptions.SSLError,
                              requests.exceptions.ConnectionError,
                              requests.exceptions.ChunkedEncodingError)):
                reset_session()
            if not quiet:
                log_print(f"[!] 请求失败（第{attempt}/{retries}次）: {type(e).__name__}: "
                          f"{str(e)[:120]}")
            interval_sleep(retry_sleep * (2 ** (attempt - 1)))   # 递增退避，别死磕同一个节奏
            continue
        try:
            j = res.json()
        except ValueError:
            if not quiet:
                log_print(f"[!] 响应非JSON（第{attempt}/{retries}次）HTTP {res.status_code}: "
                          f"{res.text[:120]}")
            interval_sleep(retry_sleep * (2 ** (attempt - 1)))
            continue
        if (j.get("message") == "Request frequency too high"
                or j.get("i18nMsg") == "message.request-frequency-too-high"):
            # 关键：这段冷却期内必须完全不发请求。被拒的请求同样计入限流窗口，
            # 边等边试会不断延长惩罚（实测每5秒探一次，124秒都没恢复）
            iv = limiter.penalize() if limiter is not None else None
            tip = f"，令牌补充放缓到 {iv:.0f}s/个" if iv else ""
            log_print(f"[!] 触发接口限流，静默 {wait} 秒后重试"
                      f"（第{attempt}/{retries}次）{tip}")
            interval_sleep(wait)
            wait *= 2
            continue
        if limiter is not None:
            limiter.reward()
        return j
    return None


def search_page(site_key: str, query: str = "*", filters: dict = None,
                from_: int = None, sort_order: str = "asc",
                facet: list = None, size: int = PAGE_SIZE):
    """调用搜索接口取一页数据（含 total 与 aggregations）。失败返回 None"""
    return fetch_json(SEARCH_API,
                      build_body(site_key, query, filters, from_, sort_order, facet, size),
                      build_headers(site_key), limiter=SEARCH_LIMITER)


def facet_buckets(site_key: str, filters: dict, dim: str):
    """取某个维度的 facet 分桶，返回 (buckets, other)：
    buckets 为 [(值, 条数), ...]，other 为落在 facet_size 之外的条数。失败返回 (None, 0)

    实测 language 各桶计数与加过滤后的 total 完全一致、other=0，是可靠的互斥分区；
    publisher/subject 则大量为空或多值，需由调用方判断可信度。"""
    j = search_page(site_key, filters=filters, facet=[dim], size=1)
    if j is None:
        return None, 0
    aggs = j.get("aggregations") or {}
    node = aggs.get(f"agg_{dim}") or aggs.get(dim) or {}
    buckets = node.get("buckets") if isinstance(node, dict) else None
    if not buckets:
        return None, 0
    out = []
    for b in buckets:
        key, cnt = b.get("key"), b.get("doc_count") or 0
        # __MISSING__ 是该字段为空的书，实测无法用它反过来过滤（total=0），跳过
        if not key or key == "__MISSING__" or not cnt:
            continue
        out.append((key, cnt))
    return out, node.get("sum_other_doc_count") or 0


def translate_title(site_key: str, book_id: str, title: str) -> str:
    """调用翻译接口将书名翻译为中文。失败/无结果返回空字符串。

    翻译失败是非致命的（回退英文原名），所以重试次数和退避都比搜索接口小得多：
    两万本书里哪怕 1% 失败，按搜索接口的 4次×5秒 也要白等一个多小时。"""
    global _translate_fail
    j = fetch_json(TRANSLATE_API,
                   {"id": book_id, "title": title, "summary": "", "lang": "zh-CN"},
                   build_headers(site_key),
                   retries=TRANSLATE_RETRY, retry_sleep=TRANSLATE_RETRY_SLEEP,
                   quiet=True)
    if j is None:
        # 单条翻译失败只累计计数，不刷屏；每 50 条汇报一次
        _translate_fail += 1
        if _translate_fail % 50 == 0:
            log_print(f"[!] 翻译接口累计失败 {_translate_fail} 条（已回退英文原名，不影响采集）")
        return ""
    if j.get("code") == 200:
        data = j.get("data") or {}
        if data.get("title"):
            return str(data["title"])
    return ""


# ==================== 站点列表（首页 HTML + JS bundle 解析） ====================

def parse_sites(js: str) -> list:
    """从 JS bundle 中解析站点列表：key、英文名、中文名、跳转链接"""
    sites = []
    # 1) zh-CN 语言包中 header 段的中文名映射
    #    压缩后的 JS 里 key 有的带引号（"main-site"）有的不带（agriculture），需两遍匹配
    zh_map = {}
    pos = js.find('"main-site":"主站点"')
    if pos >= 0:
        start = js.rfind('header:{', 0, pos)
        end = js.find('}', pos)
        if 0 <= start < end:
            seg = js[start:end]
            for k, v in re.findall(r'"([a-z0-9-]+)":"([^"]*)"', seg):
                zh_map[k] = v
            for k, v in re.findall(r'[{,]([a-z0-9-]+):"([^"]*)"', seg):
                zh_map[k] = v
    # 2) 站点对象：route:"/home/{key}" 向前找 name / i18nKey
    seen = set()
    for m in re.finditer(r'route:"/home/([a-z]+)"', js):
        key = m.group(1)
        if key in seen:
            continue
        seg = js[max(0, m.start() - 900):m.start()]
        names = re.findall(r'name:"([^"]+)"', seg)
        i18n = re.findall(r'i18nKey:"header\.([^"]+)"', seg)
        if not names:
            continue
        seen.add(key)
        sites.append({
            "key": key,
            "name": names[-1],
            "name_zh": zh_map.get(i18n[-1], "") if i18n else "",
            "home_url": HOME_PAGE_TPL.format(key=key),
            "search_url": SEARCH_PAGE_TPL.format(key=key),
        })
    return sites


def fallback_sites() -> list:
    return [{"key": k, "name": en, "name_zh": zh,
             "home_url": HOME_PAGE_TPL.format(key=k),
             "search_url": SEARCH_PAGE_TPL.format(key=k)}
            for k, en, zh in FALLBACK_SITES]


def fetch_site_list() -> list:
    """访问首页获取完整 HTML，再解析 JS bundle 得到 20 个站点。失败回退内置列表"""
    headers = {"User-Agent": UA, "Accept-Language": "zh-CN,zh;q=0.9"}
    try:
        res = get_session().get(HOME_URL, headers=headers, timeout=30)
        res.raise_for_status()
        html = res.text
        log_print(f"[*] 首页HTML获取成功: {HOME_URL}（长度 {len(html)}）")
        m = re.search(r'src="(/assets/index-[^"]+\.js)"', html)
        if not m:
            log_print("[!] 首页HTML中未找到 JS bundle 引用，使用内置站点列表")
            return fallback_sites()
        js_url = urljoin(HOME_URL, m.group(1))
        res2 = get_session().get(js_url, headers=headers, timeout=60)
        res2.raise_for_status()
        js = res2.text
        log_print(f"[*] JS bundle 获取成功: {js_url}（长度 {len(js)}）")
        sites = parse_sites(js)
        if not sites:
            log_print("[!] JS bundle 中未解析出站点，使用内置站点列表")
            return fallback_sites()
        log_print(f"[*] 已解析出 {len(sites)} 个站点:")
        for s in sites:
            zh = f"（{s['name_zh']}）" if s["name_zh"] else ""
            log_print(f"    {s['name']}{zh} [{s['key']}] -> {s['home_url']}")
        return sites
    except Exception as e:
        log_print(f"[!] 站点列表获取失败: {e}，使用内置站点列表")
        return fallback_sites()


# ==================== 字段处理 ====================

def cover_url(book_id: str) -> str:
    """按官网规律由书籍ID拼接封面链接：
    wplbn9000184970 -> https://download.worldlibrary.ai/wplbn/900/018/497/0/{id}/{id}_250.png"""
    digits = book_id.replace("wplbn", "")
    if len(digits) != 10 or not digits.isdigit():
        return ""
    return (f"https://download.worldlibrary.ai/wplbn/"
            f"{digits[0:3]}/{digits[3:6]}/{digits[6:9]}/{digits[9:]}/"
            f"{book_id}/{book_id}_250.png")


def author_str(src: dict) -> str:
    author = src.get("author")
    if isinstance(author, list):
        return ", ".join(str(x) for x in author)
    return str(author) if author else ""


def has_cjk(text: str) -> bool:
    return bool(re.search(r'[一-鿿]', text))


def book_to_row(site_key: str, book_id: str, src: dict, title_zh: str) -> list:
    """单条书籍数据 → 模板一行（ISBN、SSN号、读秀号接口无字段，留空）"""
    title = src.get("title") or ""
    return [book_id,
            title_zh or title,
            author_str(src),
            src.get("publisher") or "",
            str(src.get("publication_year")) if src.get("publication_year") is not None else "",
            "",
            str(src.get("page")) if src.get("page") is not None else "",
            cover_url(book_id),
            BOOK_URL_TPL.format(key=site_key, book_id=book_id),
            "",
            ""]


# ==================== Excel 输出 + 跨运行去重 ====================

class BookStore:
    """每个站点一个"当前"Excel（output/worldlibrary_{站点key}.xlsx）：
    - 行数达到 MAX_ROWS_PER_FILE 时自动归档为 worldlibrary_{key}_{序号}.xlsx 并另开新文件
    - 跨文件去重用纯文本 worldlibrary_{key}_ids.txt（每行一个书籍ID，非数据库）：
      启动时流式读入内存；文件不存在时从当前Excel回填生成
    - 每累计 500 条落盘一次，落盘成功后同步把新ID追加到 ids.txt"""

    def __init__(self, site_key: str, flush_every: int = FLUSH_ROWS):
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        self.site_key = site_key
        self.flush_every = flush_every
        self.path = os.path.join(OUTPUT_DIR, f"worldlibrary_{site_key}.xlsx")
        self.ids_path = os.path.join(OUTPUT_DIR, f"worldlibrary_{site_key}_ids.txt")
        self.seen_ids = set()       # 已入库书籍ID（内存中以整数/字符串存储）
        self.pending_ids = []       # 本次新增待登记到 ids.txt 的ID

        header_ok = self._check_existing_header()
        # txt 注册表优先；无有效Excel时只读txt，避免把损坏文件内容当ID
        self._load_seen_ids(backfill_ok=header_ok and os.path.exists(self.path))
        if os.path.exists(self.path) and not header_ok:
            bad = self.path
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.path = os.path.join(OUTPUT_DIR, f"worldlibrary_{site_key}_{stamp}.xlsx")
            try:
                os.rename(bad, os.path.join(OUTPUT_DIR,
                                            f"worldlibrary_{site_key}_bad_{stamp}.xlsx"))
            except OSError:
                pass
            log_print(f"[!] 已有文件 {os.path.basename(bad)} 表头不匹配或打开失败，"
                      f"已改名备份，本次新建 {os.path.basename(self.path)}")
        if header_ok and os.path.exists(self.path):
            self.wb = openpyxl.load_workbook(self.path)   # 续写已有文件
            self.ws = self.wb.active
        else:
            self._create_new_workbook()

        self.row_count = self.ws.max_row - 1
        self.pending = 0
        self.added = 0      # 本次运行新增
        self.dup = 0        # 重复跳过（含历史已采集）
        self.non_book = 0   # 非 book 类型跳过
        if self.row_count >= MAX_ROWS_PER_FILE:
            self._rotate()  # 已有文件已超上限：先归档再续写

    @staticmethod
    def _id_key(book_id: str):
        """书籍ID的内存表示：wplbn+10位数字 → 整数（大幅省内存），其它格式保持字符串"""
        s = str(book_id).strip()
        if s.startswith("wplbn") and len(s) == 15 and s[5:].isdigit():
            return int(s[5:])
        return s

    def _load_seen_ids(self, backfill_ok: bool = True):
        """启动时加载历史ID：优先读 ids.txt（快）；txt 不存在/为空且Excel有效时，
        从当前Excel流式读取并把原始ID回填到txt"""
        if os.path.exists(self.ids_path) and os.path.getsize(self.ids_path) > 0:
            try:
                with open(self.ids_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            self.seen_ids.add(self._id_key(line))
                log_print(f"[*] 已从 {os.path.basename(self.ids_path)} "
                          f"加载 {len(self.seen_ids)} 条历史ID，本次续写去重")
                return
            except Exception as e:
                log_print(f"[!] 读取 {self.ids_path} 失败: {e}")
        if backfill_ok:
            # txt 丢失/为空时：扫描当前Excel + 所有归档/历史Excel（含旧版本带时间戳的文件，
            # 排除表头损坏的 _bad_ 备份），重建去重注册
            files = [self.path]
            for p in sorted(glob.glob(os.path.join(OUTPUT_DIR,
                                                   f"worldlibrary_{self.site_key}_*.xlsx"))):
                if p != self.path and "_bad_" not in os.path.basename(p):
                    files.append(p)
            raw_ids = []
            for fp in files:
                if not os.path.exists(fp):
                    continue
                try:
                    wb_ro = openpyxl.load_workbook(fp, read_only=True)
                    try:
                        rows = wb_ro.active.iter_rows(values_only=True)
                        next(rows, None)    # 跳过表头
                        for row in rows:
                            if row and row[0]:
                                raw_ids.append(str(row[0]).strip())
                    finally:
                        wb_ro.close()
                except Exception as e:
                    log_print(f"[!] 读取 {os.path.basename(fp)} 历史ID失败: {e}")
            for rid in raw_ids:
                self.seen_ids.add(self._id_key(rid))
            if raw_ids:
                with open(self.ids_path, "w", encoding="utf-8") as f:
                    for rid in raw_ids:
                        f.write(rid + "\n")
            log_print(f"[*] 已从站点Excel文件加载 {len(self.seen_ids)} 条历史ID，"
                      f"并生成去重文件 {os.path.basename(self.ids_path)}")

    def _check_existing_header(self) -> bool:
        """已有Excel表头与模板一致返回 True；文件不存在也返回 True（视为新建）"""
        if not os.path.exists(self.path):
            return True
        try:
            wb_ro = openpyxl.load_workbook(self.path, read_only=True)
            try:
                rows = wb_ro.active.iter_rows(values_only=True)
                header = next(rows, None)
                return list(header) == EXCEL_COLUMNS
            finally:
                wb_ro.close()
        except Exception:
            return False

    def _create_new_workbook(self):
        """新建带表头样式的空工作簿"""
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "书籍数据"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        col_widths = [25, 45, 30, 30, 15, 20, 12, 60, 55, 15, 20]
        for i, width in enumerate(col_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            self.ws.column_dimensions[col_letter].width = width

    def _rotate(self):
        """当前文件达到行数上限：归档为编号文件，另开新文件继续写"""
        seq = 1
        while os.path.exists(os.path.join(OUTPUT_DIR,
                                          f"worldlibrary_{self.site_key}_{seq:04d}.xlsx")):
            seq += 1
        arch = os.path.join(OUTPUT_DIR, f"worldlibrary_{self.site_key}_{seq:04d}.xlsx")
        try:
            self.wb.save(self.path)
            os.rename(self.path, arch)
            log_print(f"[滚动] {os.path.basename(self.path)} 已达 {MAX_ROWS_PER_FILE} 行上限，"
                      f"归档为 {os.path.basename(arch)}，另开新文件继续写入")
        except OSError as e:
            log_print(f"[!] 归档失败（{e}），继续写入当前文件")
            return
        self._create_new_workbook()
        self.row_count = 0
        self.pending = 0

    def add_book(self, item: dict, site_key: str):
        """单条 hit 入库：只取 book、翻译书名、按ID去重。"""
        src = item.get("source") or {}
        if src.get("document_type") != "book":
            self.non_book += 1
            return
        book_id = str(item.get("id") or "").strip()
        if not book_id:
            return
        if self._id_key(book_id) in self.seen_ids:
            self.dup += 1
            return
        self.seen_ids.add(self._id_key(book_id))
        self.pending_ids.append(book_id)
        title = src.get("title") or ""
        title_zh = ""
        if title and not has_cjk(title):
            title_zh = translate_title(site_key, book_id, title)
            interval_sleep(TRANSLATE_INTERVAL)
        self.append(book_to_row(site_key, book_id, src, title_zh))
        self.added += 1

    def append(self, row: list):
        self.ws.append(row)
        for cell in self.ws[self.ws.max_row]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        self.row_count += 1
        self.pending += 1
        if self.pending >= self.flush_every or self.row_count >= MAX_ROWS_PER_FILE:
            self.save()
            self.pending = 0
            if self.row_count >= MAX_ROWS_PER_FILE:
                self._rotate()
            else:
                log_print(f"[落盘] 已保存 {self.row_count} 条")

    def _flush_ids(self):
        """把待登记的新ID追加写入 ids.txt（Excel 落盘成功后才调用，保证不登记未持久化的ID）"""
        if self.pending_ids:
            with open(self.ids_path, "a", encoding="utf-8") as f:
                for bid in self.pending_ids:
                    f.write(bid + "\n")
            self.pending_ids = []

    def save(self):
        try:
            self.wb.save(self.path)
        except PermissionError:
            log_print(f"[!] 保存 Excel 失败：文件被占用（请关闭 Excel/WPS 中打开的同名文件），"
                      f"稍后下次落盘会自动重试")
            return
        except Exception as e:
            log_print(f"[!] 保存 Excel 失败: {e}")
            return
        self._flush_ids()

    def close(self):
        self.save()


# ==================== 采集主流程 ====================

def process_hits(site_key: str, hits: list, store: BookStore):
    for item in hits:
        check_stop()
        store.add_book(item, site_key)


def paginate_chunk(site_key: str, store: BookStore, filters: dict = None,
                   label: str = "", sort_order: str = "asc",
                   first_page: dict = None, want: int = None) -> bool:
    """沿一个排序方向翻页采集。want 指定本趟要取的条数（默认取到 total）。
    单趟最多够到 WINDOW_LIMIT 条——接口 from 硬上限 9980，再往后返回空。"""
    arrow = "正序" if sort_order == "asc" else "倒序"
    if first_page is None:
        first_page = search_page(site_key, filters=filters, sort_order=sort_order)
        if first_page is None:
            log_print(f"[!] {label}[{arrow}] 首次请求失败，跳过")
            return False
    total = first_page.get("total") or 0
    if total == 0:
        return True
    take = min(total, WINDOW_LIMIT) if want is None else min(want, total, WINDOW_LIMIT)
    pages = math.ceil(take / PAGE_SIZE)
    process_hits(site_key, first_page.get("hits") or [], store)
    for pg in range(2, pages + 1):
        check_stop()
        from_ = (pg - 1) * PAGE_SIZE
        if from_ > MAX_FROM:
            break
        j = search_page(site_key, filters=filters, from_=from_, sort_order=sort_order)
        if j is None:
            log_print(f"[!] {label}[{arrow}] 第 {pg}/{pages} 页重试仍失败，"
                      f"该趟中断（已采数据已保存）")
            return False
        process_hits(site_key, j.get("hits") or [], store)
        if pg % 20 == 0 or pg == pages:
            log_print(f"[采集] {label}[{arrow}] 第 {pg}/{pages} 页 "
                      f"累计新增{store.added}条 重复{store.dup}条")
    return True


def collect_window(site_key: str, filters: dict, store: BookStore, label: str,
                   page1: dict = None) -> bool:
    """采集一个总量已知的查询范围（应 <= FLIP_LIMIT）：
    - <=10000：正序一趟走完
    - 10000~20000：正序取前 10000 + 倒序取剩下的，两趟拼成全量
      （单方向够不到第 10000 条以后，倒序正好补上尾部）"""
    if page1 is None:
        page1 = search_page(site_key, filters=filters, sort_order="asc")
        if page1 is None:
            log_print(f"[!] {label} 首次请求失败，跳过")
            return False
    total = page1.get("total") or 0
    if total == 0:
        log_print(f"[*] {label}: 无数据，跳过")
        return True
    if total <= WINDOW_LIMIT:
        log_print(f"[*] {label}: 共 {total} 本，{math.ceil(total / PAGE_SIZE)} 页，正序翻页采集")
        return paginate_chunk(site_key, store, filters, label, "asc", page1)
    # 拼接处留 SEAM_MARGIN 条重叠：同名书排序并列时顺序不稳定，重叠靠ID去重兜掉
    rest = min(total - WINDOW_LIMIT + SEAM_MARGIN, WINDOW_LIMIT)
    lost = total - WINDOW_LIMIT - rest
    tail = f"，中间 {lost} 本够不到" if lost > 0 else ""
    log_print(f"[*] {label}: 共 {total} 本，超过单向上限（{WINDOW_LIMIT}）；"
              f"正序取前 {WINDOW_LIMIT} 条 + 倒序取后 {rest} 条拼接{tail}")
    ok = paginate_chunk(site_key, store, filters, label, "asc", page1)
    ok &= paginate_chunk(site_key, store, filters, label, "desc", want=rest)
    return ok


def collect_partition(site_key: str, filters: dict, store: BookStore, label: str,
                      dim_idx: int = 0) -> bool:
    """完整采集一个查询范围：
    - 总量 <= 20000：正倒序两趟直接拿全
    - 超过：按 PARTITION_DIMS 里的 facet 维度切成互斥子分区，逐个递归

    facet 分桶是真正的互斥分区（实测各桶计数与加过滤后的 total 完全一致），
    这点和已废弃的 title:[lo TO hi] 词条范围分片有本质区别。"""
    filters = dict(filters or {})
    page1 = search_page(site_key, filters=filters, sort_order="asc")
    if page1 is None:
        log_print(f"[!] {label} 首次请求失败，跳过")
        return False
    total = page1.get("total") or 0
    if total == 0:
        log_print(f"[*] {label}: 无数据，跳过")
        return True
    if total <= FLIP_LIMIT:
        return collect_window(site_key, filters, store, label, page1)

    # 挑分区维度。两条判据：
    # 1) 分桶必须"可信"：other==0（没有 facet_size 之外的漏桶）且 covered<=total
    #    （covered>total 说明该字段是多值的，如 subject 一本书挂多个主题，
    #    分桶计数在重复计算，根本无法用来判断覆盖了多少本书）
    # 2) 覆盖数必须超过"直接正倒序拼接"能拿到的量，否则分区反而更亏
    #    （如某年 english 21893 本，按 publisher 只覆盖 7833 本，不如拼接拿 20000）
    baseline = min(total, FLIP_LIMIT)
    best = None                 # (覆盖数, 维度, 分桶, 维度下标)
    for i in range(dim_idx, len(PARTITION_DIMS)):
        check_stop()
        dim = PARTITION_DIMS[i]
        if filters.get(dim):    # 该维度已被上层切过
            continue
        buckets, other = facet_buckets(site_key, filters, dim)
        if not buckets or len(buckets) < 2:
            continue
        covered = sum(c for _, c in buckets)
        if other or covered > total:
            reason = (f"还有 {other} 条落在 {FACET_SIZE} 个桶之外" if other
                      else f"分桶合计 {covered} 超过总数，是多值字段")
            log_print(f"[分区] {label}: {dim} 不可用（{reason}）")
            continue
        log_print(f"[分区] {label}: {dim} 可切 {len(buckets)} 桶，覆盖 {covered}/{total} 本")
        if best is None or covered > best[0]:
            best = (covered, dim, buckets, i)
        if covered >= total:    # 已完全覆盖，不必再看其它维度
            break

    if best and best[0] > baseline:
        covered, dim, buckets, i = best
        log_print(f"[分区] {label}: 共 {total} 本，超过 {FLIP_LIMIT} 条上限，"
                  f"按 {dim} 切成 {len(buckets)} 个子分区（覆盖 {covered}/{total} 本）")
        if covered < total:
            log_print(f"[!] {label}: {dim} 分桶差 {total - covered} 本未覆盖"
                      f"（该字段为空的书接口无法反查），这部分会漏采")
        ok = True
        for val, cnt in buckets:
            check_stop()
            sub = dict(filters)
            sub[dim] = [val]
            ok &= collect_partition(site_key, sub, store,
                                    f"{label}/{dim}={val}({cnt})", i + 1)
        return ok

    log_print(f"[!] {label}: 共 {total} 本，没有可信且更划算的分区维度"
              f"（最好的只覆盖 {best[0] if best else 0} 本），"
              f"改为正倒序拼接取约 {baseline} 本，其余 {total - baseline} 本采不到")
    return collect_window(site_key, filters, store, label, page1)


def _collect_query(site_key: str, filters: dict, store: BookStore, label: str) -> bool:
    """对"站点+年份(可选)"的一个查询范围完整采集"""
    return collect_partition(site_key, filters, store, label)


def collect_site(site: dict, year_start: int = None, year_end: int = None):
    """采集一个站点。year_start/year_end 为 None 表示全站采集；
    否则按年份从 year_start 到 year_end 一年一年请求"""
    global _translate_fail
    _translate_fail = 0
    reset_session()     # 每次任务用全新会话，避免复用上次残留的坏连接
    key = site["key"]
    zh = f"（{site['name_zh']}）" if site.get("name_zh") else ""
    year_info = (f"  年份范围: {year_start} ~ {year_end}（按年逐次请求）"
                 if year_start is not None else "  年份范围: 不限（全站采集）")
    log_print("\n" + "=" * 65)
    log_print(f"  📚 开始采集站点: {site['name']}{zh} [{key}]")
    log_print(f"  搜索接口: {SEARCH_API}")
    log_print(f"  Referer: {site['search_url']}")
    log_print(year_info)
    log_print(f"  每页 {PAGE_SIZE} 条  节流: 每 {SEARCH_BURST_PERIOD:.0f} 秒最多 "
              f"{SEARCH_BURST} 次  只取 document_type=book")
    log_print("=" * 65 + "\n")

    store = BookStore(key)
    try:
        if year_start is None:
            _collect_query(key, None, store, site['name'])
        else:
            for y in range(year_start, year_end + 1):
                check_stop()
                f = {"publication_year_from": y, "publication_year_to": y}
                log_print(f"\n===== 年份 {y} =====")
                _collect_query(key, f, store, f"{y}年")
    finally:
        store.close()

    log_print("\n" + "=" * 65)
    log_print(f"[+] 采集结束：本次新增 {store.added} 条，跳过重复 {store.dup} 条，"
              f"非book类型 {store.non_book} 条，文件累计 {store.row_count} 条")
    if _translate_fail:
        log_print(f"[+] 其中 {_translate_fail} 条书名翻译失败，已回退英文原名")
    log_print(f"[+] 输出文件: {os.path.abspath(store.path)}")
    log_print("=" * 65 + "\n")


# ==================== GUI 界面 ====================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("worldlibrary.ai 电子书信息采集工具")
        self.root.geometry("860x680")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        input_frame = ttk.Frame(root, padding=10)
        input_frame.pack(fill=tk.X)

        ttk.Label(input_frame, text="站点名称:").grid(row=0, column=0, sticky=tk.W, pady=3)
        self.site_var = tk.StringVar()
        self.site_combo = ttk.Combobox(input_frame, textvariable=self.site_var,
                                       state="readonly", width=52)
        self.site_combo.grid(row=0, column=1, padx=5)
        self.refresh_btn = ttk.Button(input_frame, text="刷新站点列表", command=self.load_sites)
        self.refresh_btn.grid(row=0, column=2, padx=5)

        ttk.Label(input_frame, text="开始年份:").grid(row=1, column=0, sticky=tk.W, pady=3)
        self.year_start_var = tk.StringVar(value="")
        ttk.Entry(input_frame, textvariable=self.year_start_var, width=8).grid(
            row=1, column=1, sticky=tk.W, padx=5)
        ttk.Label(input_frame, text="结束年份:").grid(row=1, column=2, sticky=tk.W, pady=3)
        self.year_end_var = tk.StringVar(value="")
        ttk.Entry(input_frame, textvariable=self.year_end_var, width=8).grid(
            row=1, column=3, sticky=tk.W, padx=5)
        ttk.Label(input_frame, text="(格式: YYYY 如 2017；留空=全站)").grid(
            row=1, column=4, sticky=tk.W, padx=5)

        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=2, column=0, columnspan=5, pady=8)
        self.start_btn = ttk.Button(btn_frame, text="开始执行", command=self.start_task)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(btn_frame, text="结束执行", command=self.stop_task,
                                   state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        ttk.Label(root, text="运行日志:", font=("Arial", 10, "bold")).pack(
            anchor=tk.W, padx=10, pady=(10, 0))
        self.log_text = scrolledtext.ScrolledText(root, height=28, state=tk.NORMAL, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.sites = []
        self.site_map = {}
        self.running = False
        self.worker_thread = None
        self.consume_log_queue()
        self.load_sites()   # 启动时后台获取站点列表

    def consume_log_queue(self):
        try:
            while True:
                msg = LOG_QUEUE.get_nowait()
                self.log_text.insert(tk.END, msg + "\n")
                self.log_text.see(tk.END)
        except Empty:
            pass
        self.root.after(50, self.consume_log_queue)

    def load_sites(self):
        if self.running:
            return
        log_print("\n[*] 正在访问首页获取站点列表...")
        self.refresh_btn.config(state=tk.DISABLED)

        def worker():
            sites = fetch_site_list()
            self.root.after(0, lambda: self._apply_sites(sites))

        threading.Thread(target=worker, daemon=True).start()

    def _apply_sites(self, sites):
        self.sites = sites
        self.site_map = {}
        labels = []
        for s in sites:
            zh = f"{s['name_zh']} " if s.get("name_zh") else ""
            label = f"{zh}({s['name']}) [{s['key']}]"
            self.site_map[label] = s
            labels.append(label)
        self.site_combo["values"] = labels
        if labels:
            self.site_combo.current(0)
        self.refresh_btn.config(state=tk.NORMAL)
        log_print(f"[*] 站点列表加载完成，共 {len(sites)} 个站点\n")

    def on_close(self):
        self.stop_task()
        self.root.destroy()

    def validate_years(self):
        """校验年份输入。合法返回 (start, end)（可为 (None, None) 表示不限），非法返回 None"""
        s = self.year_start_var.get().strip()
        e = self.year_end_var.get().strip()
        if not s and not e:
            return None, None
        if not s or not e:
            log_print("[!] 错误：开始年份与结束年份需同时填写（都留空表示不限年份）")
            return None
        if not re.match(r"^\d{4}$", s) or not re.match(r"^\d{4}$", e):
            log_print("[!] 错误：年份格式应为 YYYY（如 2017）")
            return None
        ys, ye = int(s), int(e)
        if ys > ye:
            log_print("[!] 错误：开始年份不能晚于结束年份")
            return None
        if ys < 0 or ye > 9999:
            log_print("[!] 错误：年份超出范围（0-9999）")
            return None
        if ye - ys > 100:
            log_print(f"[*] 提示：年份跨度 {ye - ys + 1} 年，将按年逐次请求，耗时较长")
        return ys, ye

    def start_task(self):
        if self.running:
            return
        site = self.site_map.get(self.site_var.get())
        if not site:
            log_print("[!] 错误：请先选择要采集的站点名称")
            return
        years = self.validate_years()
        if years is None:
            return
        self.running = True
        _stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.refresh_btn.config(state=tk.DISABLED)

        self.worker_thread = threading.Thread(
            target=self.run_loop, args=(site, years[0], years[1]), daemon=True)
        self.worker_thread.start()

    def stop_task(self):
        if not self.running:
            return
        self.running = False
        _stop_event.set()
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        log_print("\n[!] 已发送结束信号，正在中断当前操作...")

    def run_loop(self, site, year_start=None, year_end=None):
        try:
            collect_site(site, year_start, year_end)
        except TaskStoppedException:
            log_print("\n[!] 任务已被用户结束\n")
        except Exception as e:
            log_print(f"[!] 执行异常: {e}")
            import traceback
            log_print(traceback.format_exc())
        finally:
            self.running = False
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.refresh_btn.config(state=tk.NORMAL)
            log_print("\n[*] 本次执行结束\n")


def main():
    root = tk.Tk()
    app = App(root)
    if PROXIES:
        log_print(f"[*] 当前使用代理: {PROXIES['https']}")
    else:
        log_print("[!] 未检测到代理，将直连访问；若网络受限，请在配置区设置 PROXY_OVERRIDE，"
                  "例如 PROXY_OVERRIDE = \"http://127.0.0.1:7890\"（端口看你的代理客户端）")
    root.mainloop()


if __name__ == "__main__":
    main()
